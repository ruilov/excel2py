import calendar
import math
from datetime import date
from datetime import datetime
from datetime import timedelta

from openpyxl.utils import cell as xl_cell


class ExcelError(Exception):
    pass


def normalize_addr(addr: str) -> str:
    return addr.replace("$", "").upper()


def cell_key(sheet_idx: int, addr: str) -> tuple[int, str]:
    return (sheet_idx, normalize_addr(addr))


def get_cell(cells: dict[tuple[int, str], object], sheet_idx: int, addr: str) -> object:
    return cells.get(cell_key(sheet_idx, addr))


def set_cell(cells: dict[tuple[int, str], object], sheet_idx: int, addr: str, value: object) -> None:
    cells[cell_key(sheet_idx, addr)] = value


def _addr_kind(addr: str) -> str:
    normalized = normalize_addr(addr)
    if ":" not in normalized:
        return "single"

    left, right = normalized.split(":", 1)
    if left.isdigit() and right.isdigit():
        return "row_range"
    if left.isalpha() and right.isalpha():
        return "col_range"
    return "rect_range"


def get_range(
    cells: dict[tuple[int, str], object],
    sheet_idx: int,
    addr: str,
    sheet_dimension: str | None = None,
) -> list[object] | list[list[object]]:
    normalized = normalize_addr(addr)
    kind = _addr_kind(normalized)

    if kind == "single":
        return [get_cell(cells, sheet_idx, normalized)]

    if kind == "rect_range":
        min_col, min_row, max_col, max_row = xl_cell.range_boundaries(normalized)
        rows: list[list[object]] = []
        for row in range(min_row, max_row + 1):
            current_row: list[object] = []
            for col in range(min_col, max_col + 1):
                coord = f"{xl_cell.get_column_letter(col)}{row}"
                current_row.append(get_cell(cells, sheet_idx, coord))
            rows.append(current_row)

        if len(rows) == 1:
            return rows[0]
        if all(len(row) == 1 for row in rows):
            return [row[0] for row in rows]
        return rows

    if kind == "row_range":
        if sheet_dimension is None:
            raise ValueError("sheet_dimension is required for row ranges")
        left, right = normalized.split(":", 1)
        min_row = int(left)
        max_row = int(right)
        if min_row > max_row:
            min_row, max_row = max_row, min_row

        min_col, _min_dim_row, max_col, _max_dim_row = xl_cell.range_boundaries(sheet_dimension)
        rows: list[list[object]] = []
        for row in range(min_row, max_row + 1):
            current_row: list[object] = []
            for col in range(min_col, max_col + 1):
                coord = f"{xl_cell.get_column_letter(col)}{row}"
                current_row.append(get_cell(cells, sheet_idx, coord))
            rows.append(current_row)
        return rows

    if kind == "col_range":
        if sheet_dimension is None:
            raise ValueError("sheet_dimension is required for column ranges")
        left, right = normalized.split(":", 1)
        min_col = xl_cell.column_index_from_string(left)
        max_col = xl_cell.column_index_from_string(right)
        if min_col > max_col:
            min_col, max_col = max_col, min_col

        _min_dim_col, min_row, _max_dim_col, max_row = xl_cell.range_boundaries(sheet_dimension)
        cols: list[list[object]] = []
        for col in range(min_col, max_col + 1):
            current_col: list[object] = []
            for row in range(min_row, max_row + 1):
                coord = f"{xl_cell.get_column_letter(col)}{row}"
                current_col.append(get_cell(cells, sheet_idx, coord))
            cols.append(current_col)
        return cols

    raise ValueError(f"Unsupported address reference: {addr}")


def resolve_dependency(
    cells: dict[tuple[int, str], object],
    dependency: list[object],
    sheet_dimensions: dict[int, str] | None = None,
) -> object:
    if len(dependency) == 2 and isinstance(dependency[0], int):
        sheet_idx = dependency[0]
        addr = str(dependency[1])
        if ":" in addr:
            sheet_dimension = None if sheet_dimensions is None else sheet_dimensions.get(sheet_idx)
            return get_range(cells, sheet_idx, addr, sheet_dimension=sheet_dimension)
        return get_cell(cells, sheet_idx, addr)

    if len(dependency) == 2 and dependency[0] == "name":
        raise NotImplementedError(f"Named range dependency not implemented: {dependency[1]}")

    if len(dependency) == 2 and dependency[0] == "ext":
        raise NotImplementedError(f"External reference dependency not implemented: {dependency[1]}")

    raise ValueError(f"Unsupported dependency shape: {dependency}")


def xl_if(condition: object, value_if_true: object, value_if_false: object = False) -> object:
    return value_if_true if bool(condition) else value_if_false


def xl_iferror(value: object, fallback: object) -> object:
    return fallback if isinstance(value, (Exception, ExcelError)) else value


def _iter_scalars(value: object):
    if isinstance(value, list):
        for item in value:
            yield from _iter_scalars(item)
        return
    yield value


def _coerce_number(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, bool):
        return 1.0 if value else 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str) and value.strip() == "":
        return 0.0
    return float(value)


def _try_coerce_number(value: object) -> tuple[float, bool]:
    try:
        return _coerce_number(value), True
    except (TypeError, ValueError):
        return 0.0, False


def xl_add(left: object, right: object) -> float:
    return _coerce_number(left) + _coerce_number(right)


def xl_sub(left: object, right: object) -> float:
    return _coerce_number(left) - _coerce_number(right)


def xl_mul(left: object, right: object) -> float:
    return _coerce_number(left) * _coerce_number(right)


def xl_div(left: object, right: object) -> float:
    return _coerce_number(left) / _coerce_number(right)


def xl_pow(left: object, right: object) -> float:
    return _coerce_number(left) ** _coerce_number(right)


def xl_pos(value: object) -> float:
    return _coerce_number(value)


def xl_neg(value: object) -> float:
    return -_coerce_number(value)


def xl_percent(value: object) -> float:
    return _coerce_number(value) / 100.0


def xl_sum(*args: object) -> float:
    total = 0.0
    for arg in args:
        for value in _iter_scalars(arg):
            total += _coerce_number(value)
    return total


def xl_min(*args: object) -> float:
    values = [_coerce_number(v) for arg in args for v in _iter_scalars(arg) if v is not None]
    if not values:
        raise ValueError("MIN requires at least one value")
    return min(values)


def xl_max(*args: object) -> float:
    values = [_coerce_number(v) for arg in args for v in _iter_scalars(arg) if v is not None]
    if not values:
        raise ValueError("MAX requires at least one value")
    return max(values)


def xl_and(*args: object) -> bool:
    return all(bool(value) for arg in args for value in _iter_scalars(arg))


def xl_or(*args: object) -> bool:
    return any(bool(value) for arg in args for value in _iter_scalars(arg))


def xl_not(value: object) -> bool:
    return not bool(value)


def xl_round(value: object, digits: object = 0) -> float:
    return round(_coerce_number(value), int(_coerce_number(digits)))


def xl_abs(value: object) -> float:
    return abs(_coerce_number(value))


def xl_ceiling(number: object, significance: object = 1) -> float:
    num = _coerce_number(number)
    sig = _coerce_number(significance)
    if sig == 0:
        return 0.0
    return math.ceil(num / sig) * sig


def xl_choose(index_num: object, *values: object) -> object:
    index = int(_coerce_number(index_num))
    if index < 1 or index > len(values):
        raise ValueError("CHOOSE index out of range")
    return values[index - 1]


def _to_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        # Excel serial date system (Windows 1900 date base).
        return (date(1899, 12, 30) + timedelta(days=int(value)))
    if isinstance(value, str):
        text = value.strip()
        try:
            return datetime.fromisoformat(text).date()
        except ValueError:
            return date.fromisoformat(text)
    raise ValueError(f"Unsupported date value for EOMONTH: {value!r}")


def xl_eomonth(start_date: object, months: object) -> str:
    base = _to_date(start_date)
    month_offset = int(_coerce_number(months))

    month_index = base.month - 1 + month_offset
    year = base.year + month_index // 12
    month = month_index % 12 + 1
    day = calendar.monthrange(year, month)[1]
    return datetime(year, month, day).isoformat()


def _coerce_criteria_value(text: str) -> object:
    cleaned = text.strip()
    if cleaned == "":
        return ""
    try:
        return float(cleaned)
    except ValueError:
        return cleaned


def _criteria_match(value: object, criteria: object) -> bool:
    if isinstance(criteria, str):
        crit = criteria.strip()
        for op in ("<>", "<=", ">=", "<", ">", "="):
            if crit.startswith(op):
                right = _coerce_criteria_value(crit[len(op):])
                return xl_compare(op, value, right)

        right = _coerce_criteria_value(crit)
        if isinstance(right, (int, float)) and isinstance(value, (int, float)):
            return float(value) == float(right)
        return value == right

    if isinstance(criteria, (int, float)) and isinstance(value, (int, float)):
        return float(value) == float(criteria)
    return value == criteria


def _to_1d(values: object) -> list[object]:
    if not isinstance(values, list):
        return [values]
    if values and isinstance(values[0], list):
        flattened = []
        for row in values:
            if isinstance(row, list):
                flattened.extend(row)
            else:
                flattened.append(row)
        return flattened
    return list(values)


def xl_sumif(range_values: object, criteria: object, sum_range: object | None = None) -> float:
    criteria_values = _to_1d(range_values)
    target_values = criteria_values if sum_range is None else _to_1d(sum_range)
    if len(criteria_values) != len(target_values):
        raise ValueError("SUMIF ranges must have the same length")

    total = 0.0
    for current, target in zip(criteria_values, target_values):
        if _criteria_match(current, criteria):
            total += _coerce_number(target)
    return total


def xl_sumifs(sum_range: object, *criteria_pairs: object) -> float:
    if len(criteria_pairs) % 2 != 0:
        raise ValueError("SUMIFS requires criteria_range/criteria pairs")

    target_values = _to_1d(sum_range)
    criteria_ranges: list[list[object]] = []
    criteria_values: list[object] = []

    for idx in range(0, len(criteria_pairs), 2):
        current_range = _to_1d(criteria_pairs[idx])
        if len(current_range) != len(target_values):
            raise ValueError("SUMIFS ranges must have the same length")
        criteria_ranges.append(current_range)
        criteria_values.append(criteria_pairs[idx + 1])

    total = 0.0
    for i, target in enumerate(target_values):
        matches = True
        for current_range, current_criteria in zip(criteria_ranges, criteria_values):
            if not _criteria_match(current_range[i], current_criteria):
                matches = False
                break
        if matches:
            total += _coerce_number(target)
    return total


def _npv(rate: float, cashflows: list[float]) -> float:
    total = 0.0
    for idx, value in enumerate(cashflows):
        total += value / ((1.0 + rate) ** idx)
    return total


def _npv_derivative(rate: float, cashflows: list[float]) -> float:
    total = 0.0
    for idx, value in enumerate(cashflows):
        if idx == 0:
            continue
        total -= idx * value / ((1.0 + rate) ** (idx + 1))
    return total


def xl_irr(values: object, guess: object = 0.1, max_iterations: int = 200, tolerance: float = 1e-10) -> float:
    cashflows = [_coerce_number(v) for v in _to_1d(values)]
    if not cashflows:
        raise ValueError("IRR requires at least one cashflow")
    if not any(v > 0 for v in cashflows) or not any(v < 0 for v in cashflows):
        raise ValueError("IRR requires both positive and negative cashflows")

    rate = float(guess)
    if rate <= -0.999999:
        rate = -0.9

    for _ in range(max_iterations):
        value = _npv(rate, cashflows)
        deriv = _npv_derivative(rate, cashflows)
        if abs(deriv) < 1e-14:
            break
        next_rate = rate - value / deriv
        if next_rate <= -0.999999 or not math.isfinite(next_rate):
            break
        if abs(next_rate - rate) <= tolerance:
            return next_rate
        rate = next_rate

    low = -0.999999
    high = 1.0
    low_value = _npv(low, cashflows)
    high_value = _npv(high, cashflows)
    while low_value * high_value > 0 and high < 1e6:
        high *= 2.0
        high_value = _npv(high, cashflows)

    if low_value * high_value > 0:
        raise ValueError("IRR did not converge")

    for _ in range(max_iterations):
        mid = (low + high) / 2.0
        mid_value = _npv(mid, cashflows)
        if abs(mid_value) <= tolerance or abs(high - low) <= tolerance:
            return mid
        if low_value * mid_value < 0:
            high = mid
            high_value = mid_value
        else:
            low = mid
            low_value = mid_value

    raise ValueError("IRR did not converge")


def xl_concat(left: object, right: object) -> str:
    def _to_text(v: object) -> str:
        if v is None:
            return ""
        if isinstance(v, bool):
            return "TRUE" if v else "FALSE"
        return str(v)

    return _to_text(left) + _to_text(right)


def xl_compare(op: str, left: object, right: object) -> bool:
    left_num, left_is_num = _try_coerce_number(left)
    right_num, right_is_num = _try_coerce_number(right)

    if left_is_num and right_is_num:
        left = left_num
        right = right_num
    else:
        if left is None:
            left = ""
        if right is None:
            right = ""

    if op == "=":
        return left == right
    if op == "<>":
        return left != right
    if op == "<":
        try:
            return left < right
        except TypeError:
            return str(left) < str(right)
    if op == ">":
        try:
            return left > right
        except TypeError:
            return str(left) > str(right)
    if op == "<=":
        try:
            return left <= right
        except TypeError:
            return str(left) <= str(right)
    if op == ">=":
        try:
            return left >= right
        except TypeError:
            return str(left) >= str(right)
    raise ValueError(f"Unsupported comparison operator: {op}")


def xl_error(code: str) -> ExcelError:
    return ExcelError(code)


def xl_index(array: list[object] | list[list[object]], row_num: int, col_num: int | None = None) -> object:
    if row_num is None:
        raise ValueError("row_num cannot be None")
    row_index = int(row_num) - 1

    if isinstance(array, list) and array and isinstance(array[0], list):
        col_index = 0 if col_num is None else int(col_num) - 1
        return array[row_index][col_index]

    if isinstance(array, list):
        return array[row_index]

    raise ValueError("Unsupported array shape for INDEX")


def xl_match(lookup_value: object, lookup_array: list[object], match_type: int = 1) -> int:
    if lookup_array and isinstance(lookup_array[0], list):
        if all(len(row) == 1 for row in lookup_array):
            lookup_array = [row[0] for row in lookup_array]
        elif len(lookup_array) == 1:
            lookup_array = lookup_array[0]
        else:
            raise ValueError("MATCH expects a 1D lookup array in v0")

    if match_type == 0:
        for idx, current in enumerate(lookup_array, start=1):
            if current == lookup_value:
                return idx
        raise ValueError("MATCH exact lookup failed")
    raise NotImplementedError("Only MATCH(..., ..., 0) is implemented in v0")


def xl_call(function_name: str, *args: object) -> object:
    upper = function_name.upper()
    if upper == "SUM":
        return xl_sum(*args)
    if upper == "SUMIF":
        if len(args) == 2:
            return xl_sumif(args[0], args[1])
        if len(args) >= 3:
            return xl_sumif(args[0], args[1], args[2])
        raise ValueError("SUMIF requires at least 2 arguments")
    if upper == "SUMIFS":
        if len(args) < 3:
            raise ValueError("SUMIFS requires at least 3 arguments")
        return xl_sumifs(args[0], *args[1:])
    if upper == "MIN":
        return xl_min(*args)
    if upper == "MAX":
        return xl_max(*args)
    if upper == "ABS":
        if not args:
            raise ValueError("ABS requires 1 argument")
        return xl_abs(args[0])
    if upper == "CEILING":
        if len(args) == 1:
            return xl_ceiling(args[0], 1)
        if len(args) >= 2:
            return xl_ceiling(args[0], args[1])
        raise ValueError("CEILING requires at least 1 argument")
    if upper == "CHOOSE":
        if len(args) < 2:
            raise ValueError("CHOOSE requires at least 2 arguments")
        return xl_choose(args[0], *args[1:])
    if upper == "EOMONTH":
        if len(args) < 2:
            raise ValueError("EOMONTH requires 2 arguments")
        return xl_eomonth(args[0], args[1])
    if upper == "IRR":
        if len(args) == 1:
            return xl_irr(args[0], 0.1)
        if len(args) >= 2:
            return xl_irr(args[0], args[1])
        raise ValueError("IRR requires at least 1 argument")
    if upper == "IF":
        if len(args) == 2:
            return xl_if(args[0], args[1], False)
        if len(args) >= 3:
            return xl_if(args[0], args[1], args[2])
        raise ValueError("IF requires at least 2 arguments")
    if upper == "IFERROR":
        if len(args) < 2:
            raise ValueError("IFERROR requires 2 arguments")
        return xl_iferror(args[0], args[1])
    if upper == "INDEX":
        if len(args) == 2:
            return xl_index(args[0], args[1])
        if len(args) >= 3:
            return xl_index(args[0], args[1], args[2])
        raise ValueError("INDEX requires at least 2 arguments")
    if upper == "MATCH":
        if len(args) == 2:
            return xl_match(args[0], args[1], 1)
        if len(args) >= 3:
            return xl_match(args[0], args[1], int(_coerce_number(args[2])))
        raise ValueError("MATCH requires at least 2 arguments")
    if upper == "AND":
        return xl_and(*args)
    if upper == "OR":
        return xl_or(*args)
    if upper == "NOT":
        if not args:
            raise ValueError("NOT requires 1 argument")
        return xl_not(args[0])
    if upper == "ROUND":
        if len(args) == 1:
            return xl_round(args[0], 0)
        if len(args) >= 2:
            return xl_round(args[0], args[1])
        raise ValueError("ROUND requires at least 1 argument")

    raise NotImplementedError(f"Excel function not implemented in v0 runtime: {function_name}")
