import argparse
import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl.formula.tokenizer import Tokenizer
from openpyxl.utils.cell import column_index_from_string
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.utils.cell import get_column_letter

from excel2py.ai_prompts import *
from excel2py.loader import *
from excel2py.planner import *
from excel2py.translator import *


QUALIFIED_REF_SPLIT = re.compile(r"^(?P<sheet>.+)!(?P<addr>.+)$")
A1_SINGLE_RE = re.compile(r"^(?P<col_abs>\$?)(?P<col>[A-Za-z]{1,3})(?P<row_abs>\$?)(?P<row>\d+)$")
A1_COL_RANGE_RE = re.compile(r"^(?P<left_abs>\$?)(?P<left>[A-Za-z]{1,3}):(?P<right_abs>\$?)(?P<right>[A-Za-z]{1,3})$")
A1_ROW_RANGE_RE = re.compile(r"^(?P<top_abs>\$?)(?P<top>\d+):(?P<bottom_abs>\$?)(?P<bottom>\d+)$")


def _ai_paths(workbook_path: Path, artifacts_root: str | Path) -> dict[str, Path]:
    workbook_key = workbook_path.stem
    workbook_dir = Path(artifacts_root) / workbook_key
    ai_dir = workbook_dir / "ai"
    prompts_dir = ai_dir / "prompts"
    cluster_prompt_dir = prompts_dir / "clusters"
    return {
        "workbook_dir": workbook_dir,
        "raw_dir": workbook_dir / "raw",
        "derived_dir": workbook_dir / "derived",
        "generated_dir": workbook_dir / "generated",
        "ai_dir": ai_dir,
        "prompts_dir": prompts_dir,
        "cluster_prompt_dir": cluster_prompt_dir,
        "metadata_path": workbook_dir / "raw" / "workbook_meta.json",
        "cells_path": workbook_dir / "raw" / "cells.jsonl",
        "formulas_path": workbook_dir / "derived" / "formulas.jsonl",
        "dependencies_path": workbook_dir / "derived" / "dependencies.jsonl",
        "calc_order_path": workbook_dir / "derived" / "calc_order.json",
        "literal_script_path": workbook_dir / "generated" / f"{workbook_key}_literal.py",
        "global_context_json": ai_dir / "global_context.json",
        "clusters_json": ai_dir / "clusters.json",
        "cluster_packets_jsonl": ai_dir / "cluster_packets.jsonl",
        "system_prompt_txt": prompts_dir / "system_prompt.txt",
        "global_prompt_txt": prompts_dir / "global_context_prompt.txt",
        "plan_prompt_txt": prompts_dir / "plan_synthesis_prompt.txt",
    }


def _normalize_addr(addr: str) -> str:
    return addr.replace("$", "").upper()


def _safe_id(value: str) -> str:
    return re.sub(r"[^a-zA-Z0-9_]", "_", value)


def _load_jsonl(path: Path) -> list[object]:
    rows = []
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            rows.append(json.loads(line))
    return rows


def _split_qualified_ref(ref_text: str) -> tuple[str | None, str | None]:
    match = QUALIFIED_REF_SPLIT.match(ref_text)
    if match is None:
        return None, None
    return match.group("sheet"), match.group("addr")


def _normalize_sheet_token(sheet_token: str) -> str:
    token = sheet_token.strip()
    if token.startswith("'") and token.endswith("'") and len(token) >= 2:
        token = token[1:-1].replace("''", "'")
    return token


def _sheet_idx_from_token(sheet_token: str, sheet_idx_by_name: dict[str, int]) -> int | None:
    normalized = _normalize_sheet_token(sheet_token)
    if normalized in sheet_idx_by_name:
        return sheet_idx_by_name[normalized]

    without_book_prefix = re.sub(r"^\[[^\]]+\]", "", normalized)
    if without_book_prefix in sheet_idx_by_name:
        return sheet_idx_by_name[without_book_prefix]

    return None


def _canonicalize_single_ref(text: str, origin_row: int, origin_col: int) -> str:
    match = A1_SINGLE_RE.fullmatch(text)
    if match is None:
        return text.upper()

    row_value = int(match.group("row"))
    col_value = column_index_from_string(match.group("col").upper())
    row_abs = match.group("row_abs") == "$"
    col_abs = match.group("col_abs") == "$"

    row_part = f"R{row_value}" if row_abs else f"R[{row_value - origin_row}]"
    col_part = f"C{col_value}" if col_abs else f"C[{col_value - origin_col}]"
    return f"{row_part}{col_part}"


def _canonicalize_col_range(text: str, origin_col: int) -> str:
    match = A1_COL_RANGE_RE.fullmatch(text)
    if match is None:
        return text.upper()

    left_col = column_index_from_string(match.group("left").upper())
    right_col = column_index_from_string(match.group("right").upper())
    left_abs = match.group("left_abs") == "$"
    right_abs = match.group("right_abs") == "$"

    left_part = f"C{left_col}" if left_abs else f"C[{left_col - origin_col}]"
    right_part = f"C{right_col}" if right_abs else f"C[{right_col - origin_col}]"
    return f"{left_part}:{right_part}"


def _canonicalize_row_range(text: str, origin_row: int) -> str:
    match = A1_ROW_RANGE_RE.fullmatch(text)
    if match is None:
        return text.upper()

    top_row = int(match.group("top"))
    bottom_row = int(match.group("bottom"))
    top_abs = match.group("top_abs") == "$"
    bottom_abs = match.group("bottom_abs") == "$"

    top_part = f"R{top_row}" if top_abs else f"R[{top_row - origin_row}]"
    bottom_part = f"R{bottom_row}" if bottom_abs else f"R[{bottom_row - origin_row}]"
    return f"{top_part}:{bottom_part}"


def _canonicalize_ref_text(
    ref_text: str,
    origin_sheet_idx: int,
    origin_row: int,
    origin_col: int,
    sheet_idx_by_name: dict[str, int],
) -> str:
    sheet_part = None
    addr_part = ref_text

    if "!" in ref_text:
        sheet_token, address_token = _split_qualified_ref(ref_text)
        if sheet_token is not None and address_token is not None:
            target_sheet_idx = _sheet_idx_from_token(sheet_token, sheet_idx_by_name)
            if target_sheet_idx is None:
                sheet_part = f"S[ext:{_normalize_sheet_token(sheet_token)}]"
            else:
                sheet_part = f"S[{target_sheet_idx - origin_sheet_idx}]"
            addr_part = address_token

    normalized = addr_part.upper()
    if ":" in normalized:
        left, right = normalized.split(":", 1)
        if A1_SINGLE_RE.fullmatch(left) and A1_SINGLE_RE.fullmatch(right):
            left_part = _canonicalize_single_ref(left, origin_row, origin_col)
            right_part = _canonicalize_single_ref(right, origin_row, origin_col)
            canonical_addr = f"{left_part}:{right_part}"
        elif A1_COL_RANGE_RE.fullmatch(normalized):
            canonical_addr = _canonicalize_col_range(normalized, origin_col)
        elif A1_ROW_RANGE_RE.fullmatch(normalized):
            canonical_addr = _canonicalize_row_range(normalized, origin_row)
        else:
            canonical_addr = normalized
    else:
        canonical_addr = _canonicalize_single_ref(normalized, origin_row, origin_col)

    if sheet_part is None:
        return canonical_addr
    return f"{sheet_part}!{canonical_addr}"


def _formula_signature(formula: str, sheet_idx: int, addr: str, sheet_idx_by_name: dict[str, int]) -> str:
    row, col = coordinate_to_tuple(_normalize_addr(addr))
    tokenizer = Tokenizer(formula)
    chunks: list[str] = []

    for token in tokenizer.items:
        value = token.value
        if token.type == "OPERAND" and token.subtype == "RANGE":
            value = _canonicalize_ref_text(
                ref_text=value,
                origin_sheet_idx=sheet_idx,
                origin_row=row,
                origin_col=col,
                sheet_idx_by_name=sheet_idx_by_name,
            )
        chunks.append(value.upper())

    return "".join(chunks)


def _range_hint_from_cells(cells: list[tuple[int, str]], sheet_names: list[str]) -> list[dict[str, object]]:
    by_sheet: dict[int, list[tuple[int, int, str]]] = defaultdict(list)
    for sheet_idx, addr in cells:
        row, col = coordinate_to_tuple(_normalize_addr(addr))
        by_sheet[sheet_idx].append((row, col, _normalize_addr(addr)))

    hints = []
    for sheet_idx, entries in by_sheet.items():
        min_row = min(row for row, _col, _addr in entries)
        max_row = max(row for row, _col, _addr in entries)
        min_col = min(col for _row, col, _addr in entries)
        max_col = max(col for _row, col, _addr in entries)
        hints.append(
            {
                "sheet_idx": sheet_idx,
                "sheet": sheet_names[sheet_idx],
                "range_hint": f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}",
                "cell_count": len(entries),
            }
        )
    return sorted(hints, key=lambda item: (item["sheet_idx"], item["range_hint"]))


def _collect_string_labels(cells_rows: list[object]) -> dict[int, dict[tuple[int, int], str]]:
    labels: dict[int, dict[tuple[int, int], str]] = defaultdict(dict)
    for row in cells_rows:
        sheet_idx, addr, _data_type, formula, value = row
        if formula is not None:
            continue
        if not isinstance(value, str):
            continue
        cleaned = value.strip()
        if cleaned == "":
            continue
        cell_row, cell_col = coordinate_to_tuple(_normalize_addr(addr))
        labels[sheet_idx][(cell_row, cell_col)] = cleaned
    return labels


def _nearby_labels(
    labels_by_sheet: dict[int, dict[tuple[int, int], str]],
    sheet_idx: int,
    addr: str,
    window_rows: int = 4,
    window_cols: int = 4,
    limit: int = 8,
) -> list[str]:
    row, col = coordinate_to_tuple(_normalize_addr(addr))
    labels = labels_by_sheet.get(sheet_idx, {})
    found: list[tuple[int, str]] = []
    for (label_row, label_col), text in labels.items():
        row_distance = abs(label_row - row)
        col_distance = abs(label_col - col)
        if row_distance > window_rows or col_distance > window_cols:
            continue
        found.append((row_distance + col_distance, text))
    found.sort(key=lambda item: item[0])
    out: list[str] = []
    for _distance, text in found:
        if text in out:
            continue
        out.append(text)
        if len(out) >= limit:
            break
    return out


def _build_global_context(
    metadata: dict[str, object],
    formula_rows: list[object],
    dependency_rows: list[object],
    calc_order_payload: dict[str, object],
    labels_by_sheet: dict[int, dict[tuple[int, int], str]],
) -> dict[str, object]:
    formula_count_by_sheet: dict[int, int] = defaultdict(int)
    for row in formula_rows:
        formula_count_by_sheet[row[0]] += 1

    dep_count_by_sheet: dict[int, int] = defaultdict(int)
    for row in dependency_rows:
        dep_count_by_sheet[row[0]] += len(row[2])

    sheet_summaries = []
    for sheet_idx, sheet_name in enumerate(metadata["sheet_names"]):
        sample_labels = list(labels_by_sheet.get(sheet_idx, {}).values())[:20]
        sheet_summaries.append(
            {
                "sheet_idx": sheet_idx,
                "sheet": sheet_name,
                "formula_cell_count": formula_count_by_sheet.get(sheet_idx, 0),
                "dependency_edge_count": dep_count_by_sheet.get(sheet_idx, 0),
                "sheet_dimension": metadata["sheet_dimensions"][sheet_name],
                "sample_labels": sample_labels,
            }
        )

    major_blocks = []
    for summary in sheet_summaries:
        if summary["formula_cell_count"] == 0:
            continue
        major_blocks.append(
            {
                "block_id": f"sheet_{summary['sheet_idx']}_main",
                "sheet": summary["sheet"],
                "range_hint": summary["sheet_dimension"],
                "label_evidence": summary["sample_labels"][:6],
                "purpose_hypothesis": "Main calculation region inferred from formula density and nearby labels.",
            }
        )

    named_ranges = [item["name"] for item in metadata.get("named_ranges", [])]
    workbook_summary = (
        "Workbook-level deterministic extraction summary used for AI refactor planning. "
        "Business intent should be inferred conservatively from labels and dependency structure."
    )

    return {
        "workbook_path": metadata["workbook_path"],
        "sheet_names": metadata["sheet_names"],
        "workbook_summary": workbook_summary,
        "sheet_summaries": sheet_summaries,
        "major_blocks": major_blocks,
        "named_ranges": named_ranges,
        "calc_order_count": len(calc_order_payload.get("calc_order", [])),
        "cycle_group_count": len(calc_order_payload.get("cycles", [])),
        "graph_stats": calc_order_payload.get("stats", {}),
    }


def _build_clusters(
    formula_rows: list[object],
    dependency_rows: list[object],
    metadata: dict[str, object],
) -> tuple[list[dict[str, object]], dict[tuple[int, str], str]]:
    sheet_idx_by_name = {name: idx for idx, name in enumerate(metadata["sheet_names"])}
    dependency_map = {(row[0], _normalize_addr(row[1])): row[2] for row in dependency_rows}

    grouped: dict[str, list[object]] = defaultdict(list)
    for row in formula_rows:
        sheet_idx = row[0]
        addr = _normalize_addr(row[1])
        formula = row[3]
        signature = _formula_signature(
            formula=formula,
            sheet_idx=sheet_idx,
            addr=addr,
            sheet_idx_by_name=sheet_idx_by_name,
        )
        grouped[signature].append([sheet_idx, addr, formula])

    sorted_groups = sorted(
        grouped.items(),
        key=lambda item: (-len(item[1]), item[1][0][0], item[1][0][1]),
    )

    clusters: list[dict[str, object]] = []
    cell_to_cluster: dict[tuple[int, str], str] = {}
    for index, (signature, members) in enumerate(sorted_groups, start=1):
        cluster_id = f"cluster_{index:04d}"
        member_cells = [(member[0], member[1]) for member in members]
        range_hints = _range_hint_from_cells(member_cells, metadata["sheet_names"])
        unique_formulas = []
        for _sheet_idx, _addr, formula in members:
            if formula not in unique_formulas:
                unique_formulas.append(formula)
            if len(unique_formulas) >= 3:
                break

        dependency_count = 0
        for sheet_idx, addr in member_cells:
            dependency_count += len(dependency_map.get((sheet_idx, addr), []))

        cluster = {
            "cluster_id": cluster_id,
            "signature": signature,
            "cell_count": len(member_cells),
            "anchor": [members[0][0], members[0][1]],
            "sample_formulas": unique_formulas,
            "range_hints": range_hints,
            "cells": [[sheet_idx, addr] for sheet_idx, addr in member_cells],
            "dependency_edge_count": dependency_count,
        }
        clusters.append(cluster)
        for key in member_cells:
            cell_to_cluster[key] = cluster_id

    return clusters, cell_to_cluster


def _build_cluster_graph(
    clusters: list[dict[str, object]],
    dependency_rows: list[object],
    cell_to_cluster: dict[tuple[int, str], str],
) -> dict[str, dict[str, set[str]]]:
    edges_out: dict[str, set[str]] = defaultdict(set)
    edges_in: dict[str, set[str]] = defaultdict(set)

    for row in dependency_rows:
        target_key = (row[0], _normalize_addr(row[1]))
        target_cluster = cell_to_cluster.get(target_key)
        if target_cluster is None:
            continue
        for dependency in row[2]:
            if not (isinstance(dependency, list) and len(dependency) == 2 and isinstance(dependency[0], int)):
                continue
            source_key = (dependency[0], _normalize_addr(str(dependency[1])))
            source_cluster = cell_to_cluster.get(source_key)
            if source_cluster is None or source_cluster == target_cluster:
                continue
            edges_out[source_cluster].add(target_cluster)
            edges_in[target_cluster].add(source_cluster)

    graph = {}
    for cluster in clusters:
        cluster_id = cluster["cluster_id"]
        graph[cluster_id] = {
            "upstream": edges_in.get(cluster_id, set()),
            "downstream": edges_out.get(cluster_id, set()),
        }
    return graph


def _build_cluster_packets(
    clusters: list[dict[str, object]],
    cluster_graph: dict[str, dict[str, set[str]]],
    labels_by_sheet: dict[int, dict[tuple[int, int], str]],
) -> list[dict[str, object]]:
    packets: list[dict[str, object]] = []
    for cluster in clusters:
        anchor_sheet_idx, anchor_addr = cluster["anchor"]
        nearby = _nearby_labels(
            labels_by_sheet=labels_by_sheet,
            sheet_idx=anchor_sheet_idx,
            addr=anchor_addr,
        )
        graph_node = cluster_graph.get(cluster["cluster_id"], {"upstream": set(), "downstream": set()})
        packets.append(
            {
                "cluster_id": cluster["cluster_id"],
                "anchor": cluster["anchor"],
                "cell_count": cluster["cell_count"],
                "range_hints": cluster["range_hints"],
                "sample_formulas": cluster["sample_formulas"],
                "dependency_edge_count": cluster["dependency_edge_count"],
                "upstream_clusters": sorted(graph_node["upstream"]),
                "downstream_clusters": sorted(graph_node["downstream"]),
                "nearby_labels": nearby,
            }
        )
    return packets


def _write_prompt_files(
    paths: dict[str, Path],
    global_context: dict[str, object],
    cluster_packets: list[dict[str, object]],
    literal_source_code: str,
    max_cluster_prompts: int,
    include_singletons: bool,
) -> dict[str, object]:
    paths["ai_dir"].mkdir(parents=True, exist_ok=True)
    paths["prompts_dir"].mkdir(parents=True, exist_ok=True)
    paths["cluster_prompt_dir"].mkdir(parents=True, exist_ok=True)

    paths["global_context_json"].write_text(
        json.dumps(global_context, indent=2, ensure_ascii=True) + "\n",
        encoding="utf-8",
    )

    paths["cluster_packets_jsonl"].write_text("", encoding="utf-8")
    with paths["cluster_packets_jsonl"].open("w", encoding="utf-8", newline="\n") as handle:
        for packet in cluster_packets:
            handle.write(json.dumps(packet, ensure_ascii=True) + "\n")

    paths["system_prompt_txt"].write_text(REFRACTOR_SYSTEM_PROMPT + "\n", encoding="utf-8")

    global_prompt = render_prompt(
        GLOBAL_CONTEXT_PROMPT_TEMPLATE,
        global_context_json=json.dumps(global_context, ensure_ascii=True, indent=2),
    )
    paths["global_prompt_txt"].write_text(global_prompt + "\n", encoding="utf-8")

    selected_packets = []
    for packet in cluster_packets:
        if not include_singletons and packet["cell_count"] <= 1:
            continue
        selected_packets.append(packet)

    selected_packets = selected_packets[:max_cluster_prompts]
    prompt_count = 0
    global_summary_packet = {
        "workbook_summary": global_context["workbook_summary"],
        "sheet_summaries": global_context["sheet_summaries"],
        "major_blocks": global_context["major_blocks"],
    }
    for packet in selected_packets:
        prompt = render_prompt(
            CLUSTER_PACKET_PROMPT_TEMPLATE,
            global_summary_json=json.dumps(global_summary_packet, ensure_ascii=True, indent=2),
            cluster_packet_json=json.dumps(packet, ensure_ascii=True, indent=2),
        )
        prompt_path = paths["cluster_prompt_dir"] / f"{packet['cluster_id']}.txt"
        prompt_path.write_text(prompt + "\n", encoding="utf-8")
        prompt_count += 1

    plan_prompt = render_prompt(
        PLAN_SYNTHESIS_PROMPT_TEMPLATE,
        global_summary_json=json.dumps(global_summary_packet, ensure_ascii=True, indent=2),
        cluster_proposals_json=json.dumps([], ensure_ascii=True, indent=2),
    )
    paths["plan_prompt_txt"].write_text(plan_prompt + "\n", encoding="utf-8")

    codegen_prompt_path = paths["prompts_dir"] / "codegen_prompt.txt"
    codegen_prompt = render_prompt(
        CODEGEN_PROMPT_TEMPLATE,
        literal_source_code=literal_source_code,
        approved_plan_json=json.dumps({"functions": [], "orchestration_order": []}, ensure_ascii=True, indent=2),
    )
    codegen_prompt_path.write_text(codegen_prompt + "\n", encoding="utf-8")

    self_check_prompt_path = paths["prompts_dir"] / "self_check_prompt.txt"
    self_check_prompt = render_prompt(
        SELF_CHECK_PROMPT_TEMPLATE,
        refactored_code="# Paste candidate refactored code here",
    )
    self_check_prompt_path.write_text(self_check_prompt + "\n", encoding="utf-8")

    return {
        "global_prompt_path": str(paths["global_prompt_txt"].resolve()),
        "cluster_prompt_count": prompt_count,
        "cluster_prompt_dir": str(paths["cluster_prompt_dir"].resolve()),
        "plan_prompt_path": str(paths["plan_prompt_txt"].resolve()),
        "codegen_prompt_path": str(codegen_prompt_path.resolve()),
        "self_check_prompt_path": str(self_check_prompt_path.resolve()),
    }


def prepare_ai_refactor_inputs(
    excel_file: str | Path = DEFAULT_WORKBOOK,
    artifacts_root: str | Path = DEFAULT_ARTIFACTS_ROOT,
    max_cluster_prompts: int = 200,
    include_singletons: bool = False,
) -> dict[str, object]:
    workbook_path = Path(excel_file)
    paths = _ai_paths(workbook_path, artifacts_root)

    export_workbook_artifacts(excel_file=workbook_path, artifacts_root=artifacts_root)
    export_formula_rows(excel_file=workbook_path, artifacts_root=artifacts_root)
    dep_result = export_formula_dependencies(excel_file=workbook_path, artifacts_root=artifacts_root)
    build_calc_order(
        excel_file=workbook_path,
        artifacts_root=artifacts_root,
        parse_diagnostics=dep_result,
    )
    emit_literal_skeleton(excel_file=workbook_path, artifacts_root=artifacts_root)

    metadata = json.loads(paths["metadata_path"].read_text(encoding="utf-8"))
    cells_rows = _load_jsonl(paths["cells_path"])
    formula_rows = _load_jsonl(paths["formulas_path"])
    dependency_rows = _load_jsonl(paths["dependencies_path"])
    calc_order_payload = json.loads(paths["calc_order_path"].read_text(encoding="utf-8"))
    literal_source_code = paths["literal_script_path"].read_text(encoding="utf-8")

    labels_by_sheet = _collect_string_labels(cells_rows)
    global_context = _build_global_context(
        metadata=metadata,
        formula_rows=formula_rows,
        dependency_rows=dependency_rows,
        calc_order_payload=calc_order_payload,
        labels_by_sheet=labels_by_sheet,
    )
    clusters, cell_to_cluster = _build_clusters(
        formula_rows=formula_rows,
        dependency_rows=dependency_rows,
        metadata=metadata,
    )
    cluster_graph = _build_cluster_graph(
        clusters=clusters,
        dependency_rows=dependency_rows,
        cell_to_cluster=cell_to_cluster,
    )
    cluster_packets = _build_cluster_packets(
        clusters=clusters,
        cluster_graph=cluster_graph,
        labels_by_sheet=labels_by_sheet,
    )

    paths["ai_dir"].mkdir(parents=True, exist_ok=True)
    paths["clusters_json"].write_text(
        json.dumps(clusters, indent=2, ensure_ascii=True) + "\n",
        encoding="utf-8",
    )

    prompt_result = _write_prompt_files(
        paths=paths,
        global_context=global_context,
        cluster_packets=cluster_packets,
        literal_source_code=literal_source_code,
        max_cluster_prompts=max_cluster_prompts,
        include_singletons=include_singletons,
    )

    repeated_cluster_count = sum(1 for cluster in clusters if cluster["cell_count"] > 1)
    return {
        "workbook_path": str(workbook_path.resolve()),
        "ai_dir": str(paths["ai_dir"].resolve()),
        "global_context_path": str(paths["global_context_json"].resolve()),
        "clusters_path": str(paths["clusters_json"].resolve()),
        "cluster_packets_path": str(paths["cluster_packets_jsonl"].resolve()),
        "cluster_count": len(clusters),
        "repeated_cluster_count": repeated_cluster_count,
        "prompt_outputs": prompt_result,
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Prepare deterministic context + prompt packets for AI-assisted refactor."
    )
    parser.add_argument(
        "excel_file",
        nargs="?",
        default=str(DEFAULT_WORKBOOK),
        help=f"Path (or base name) of workbook. Default: {DEFAULT_WORKBOOK.name}",
    )
    parser.add_argument(
        "--artifacts-root",
        default=str(DEFAULT_ARTIFACTS_ROOT),
        help=f"Root folder for generated artifacts. Default: {DEFAULT_ARTIFACTS_ROOT}",
    )
    parser.add_argument(
        "--max-cluster-prompts",
        type=int,
        default=200,
        help="Maximum number of per-cluster prompt files to emit.",
    )
    parser.add_argument(
        "--include-singletons",
        action="store_true",
        help="Include single-cell clusters when emitting per-cluster prompts.",
    )
    args = parser.parse_args()

    result = prepare_ai_refactor_inputs(
        excel_file=args.excel_file,
        artifacts_root=args.artifacts_root,
        max_cluster_prompts=args.max_cluster_prompts,
        include_singletons=args.include_singletons,
    )

    print(f"Workbook: {result['workbook_path']}")
    print(f"AI output dir: {result['ai_dir']}")
    print(f"Global context: {result['global_context_path']}")
    print(f"Clusters JSON: {result['clusters_path']}")
    print(f"Cluster packets JSONL: {result['cluster_packets_path']}")
    print(f"Clusters: {result['cluster_count']}")
    print(f"Repeated clusters: {result['repeated_cluster_count']}")
    print(f"Global prompt: {result['prompt_outputs']['global_prompt_path']}")
    print(f"Cluster prompts: {result['prompt_outputs']['cluster_prompt_count']}")
    print(f"Plan prompt: {result['prompt_outputs']['plan_prompt_path']}")
    print(f"Codegen prompt: {result['prompt_outputs']['codegen_prompt_path']}")
    print(f"Self-check prompt: {result['prompt_outputs']['self_check_prompt_path']}")


if __name__ == "__main__":
    main()
