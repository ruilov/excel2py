import json
import re
from datetime import *
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter

DEFAULT_WORKBOOK = Path("excel_model.xlsx")
DEFAULT_ARTIFACTS_ROOT = Path("artifacts")
CELL_RECORD_FIELDS = ["sheet_idx", "addr", "data_type", "formula", "value"]
NAME_TOKEN_PATTERN = re.compile(r"[A-Za-z_\\][A-Za-z0-9_.\\]*")


def list_sheet_names(excel_file: str | Path = DEFAULT_WORKBOOK) -> list[str]:
    workbook_path = Path(excel_file)
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def read_workbook_metadata(excel_file: str | Path = DEFAULT_WORKBOOK) -> dict[str, object]:
    return _read_workbook_metadata(excel_file)


def _read_workbook_metadata(
    excel_file: str | Path = DEFAULT_WORKBOOK,
    used_named_ranges: set[str] | None = None,
) -> dict[str, object]:
    workbook_path = Path(excel_file)
    workbook_formulas = load_workbook(workbook_path, data_only=False, read_only=False)
    workbook_values = load_workbook(workbook_path, data_only=True, read_only=False)
    try:
        sheet_dimensions = {}
        for sheet_name in workbook_formulas.sheetnames:
            sheet_formulas = workbook_formulas[sheet_name]
            sheet_values = workbook_values[sheet_name]
            min_col, min_row, max_col, max_row = _sheet_iteration_bounds(
                sheet_formulas,
                sheet_values,
            )
            sheet_dimensions[sheet_name] = (
                f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
            )

        named_ranges = []
        for range_name, defined_name in workbook_formulas.defined_names.items():
            if used_named_ranges is not None and range_name not in used_named_ranges:
                continue

            entries = defined_name if isinstance(defined_name, list) else [defined_name]
            for entry in entries:
                destinations = []
                try:
                    for sheet_name, cell_range in entry.destinations:
                        destinations.append({"sheet": sheet_name, "range": cell_range})
                except (TypeError, ValueError, AttributeError):
                    destinations = []

                named_ranges.append(
                    {
                        "name": range_name,
                        "value": getattr(entry, "attr_text", None),
                        "destinations": destinations,
                    }
                )

        return {
            "workbook_path": str(workbook_path.resolve()),
            "sheet_names": list(workbook_formulas.sheetnames),
            "sheet_dimensions": sheet_dimensions,
            "named_ranges": named_ranges,
        }
    finally:
        workbook_formulas.close()
        workbook_values.close()


def _find_defined_names_in_text(
    text: object,
    defined_name_lookup: dict[str, str],
) -> set[str]:
    if not isinstance(text, str):
        return set()

    candidates = set()
    for token in NAME_TOKEN_PATTERN.findall(text):
        range_name = defined_name_lookup.get(token.upper())
        if range_name is not None:
            candidates.add(range_name)
    return candidates


def _to_jsonable(value: object) -> object:
    if value is None or isinstance(value, (bool, int, float, str)):
        return value
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, timedelta):
        return value.total_seconds()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    return str(value)


def _artifact_paths(workbook_path: Path, artifacts_root: str | Path) -> dict[str, Path]:
    workbook_key = workbook_path.stem
    raw_dir = Path(artifacts_root) / workbook_key / "raw"
    return {
        "artifact_dir": raw_dir,
        "manifest_path": raw_dir / "manifest.json",
        "metadata_path": raw_dir / "workbook_meta.json",
        "cells_path": raw_dir / "cells.jsonl",
    }


def _sheet_iteration_bounds(
    sheet_formulas,
    sheet_values,
) -> tuple[int, int, int, int]:
    min_row = None
    min_col = None
    max_row = 0
    max_col = 0

    coordinates = set(sheet_formulas._cells.keys()) | set(sheet_values._cells.keys())
    for row, col in coordinates:
        formula_cell = sheet_formulas._cells.get((row, col))
        value_cell = sheet_values._cells.get((row, col))

        has_formula = (
            formula_cell is not None
            and formula_cell.data_type == "f"
            and formula_cell.value is not None
        )
        has_value = value_cell is not None and value_cell.value is not None
        if not has_formula and not has_value:
            continue

        if min_row is None or row < min_row:
            min_row = row
        if min_col is None or col < min_col:
            min_col = col
        if row > max_row:
            max_row = row
        if col > max_col:
            max_col = col

    if min_row is None or min_col is None:
        return 1, 1, 1, 1

    return min_col, min_row, max_col, max_row


def export_workbook_artifacts(
    excel_file: str | Path = DEFAULT_WORKBOOK,
    artifacts_root: str | Path = DEFAULT_ARTIFACTS_ROOT,
) -> dict[str, object]:
    workbook_path = Path(excel_file)
    paths = _artifact_paths(workbook_path, artifacts_root)
    paths["artifact_dir"].mkdir(parents=True, exist_ok=True)

    workbook_values = load_workbook(workbook_path, data_only=True, read_only=False)
    workbook_formulas = load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        cell_count = 0
        formula_cell_count = 0
        used_named_ranges = set()
        defined_name_lookup = {
            range_name.upper(): range_name for range_name in workbook_formulas.defined_names.keys()
        }

        with paths["cells_path"].open("w", encoding="utf-8", newline="\n") as handle:
            for sheet_idx, sheet_name in enumerate(workbook_formulas.sheetnames):
                sheet_formulas = workbook_formulas[sheet_name]
                sheet_values = workbook_values[sheet_name]
                min_col, min_row, max_col, max_row = _sheet_iteration_bounds(
                    sheet_formulas,
                    sheet_values,
                )
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        formula_cell = sheet_formulas.cell(row=row, column=col)
                        value_cell = sheet_values.cell(row=row, column=col)
                        is_formula = formula_cell.data_type == "f"
                        has_value = value_cell.value is not None
                        has_formula = is_formula and formula_cell.value is not None

                        if not has_formula and not has_value:
                            continue

                        if is_formula:
                            formula_cell_count += 1
                            used_named_ranges.update(
                                _find_defined_names_in_text(formula_cell.value, defined_name_lookup)
                            )
                        elif has_value:
                            used_named_ranges.update(
                                _find_defined_names_in_text(value_cell.value, defined_name_lookup)
                            )

                        record = [
                            sheet_idx,
                            formula_cell.coordinate,
                            formula_cell.data_type,
                            _to_jsonable(formula_cell.value) if is_formula else None,
                            _to_jsonable(value_cell.value),
                        ]
                        handle.write(json.dumps(record, ensure_ascii=True) + "\n")
                        cell_count += 1

        metadata = _read_workbook_metadata(
            workbook_path, used_named_ranges=used_named_ranges
        )
        metadata["cell_count"] = cell_count
        metadata["formula_cell_count"] = formula_cell_count
        metadata["cells_jsonl_record_fields"] = CELL_RECORD_FIELDS
        metadata["cells_jsonl_sheet_lookup"] = "sheet_names"
        metadata["cells_jsonl_sheet_index_base"] = 0
        paths["metadata_path"].write_text(
            json.dumps(metadata, indent=2, ensure_ascii=True) + "\n",
            encoding="utf-8",
        )

        manifest = {
            "exported_at_utc": datetime.now(timezone.utc).isoformat(),
            "source_workbook": str(workbook_path.resolve()),
            "artifacts": {
                "workbook_meta_json": paths["metadata_path"].name,
                "cells_jsonl": paths["cells_path"].name,
            },
            "cells_jsonl_record_fields": CELL_RECORD_FIELDS,
            "cells_jsonl_sheet_lookup": "sheet_names",
            "cells_jsonl_sheet_index_base": 0,
            "stats": {
                "sheet_count": len(workbook_formulas.sheetnames),
                "cell_count": cell_count,
                "formula_cell_count": formula_cell_count,
            },
        }
        paths["manifest_path"].write_text(
            json.dumps(manifest, indent=2, ensure_ascii=True) + "\n",
            encoding="utf-8",
        )

        return {
            "workbook_path": str(workbook_path.resolve()),
            "artifact_dir": str(paths["artifact_dir"].resolve()),
            "manifest_path": str(paths["manifest_path"].resolve()),
            "metadata_path": str(paths["metadata_path"].resolve()),
            "cells_path": str(paths["cells_path"].resolve()),
            "sheet_count": len(workbook_formulas.sheetnames),
            "cell_count": cell_count,
            "formula_cell_count": formula_cell_count,
        }
    finally:
        workbook_formulas.close()
        workbook_values.close()
